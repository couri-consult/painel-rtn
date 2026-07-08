import openpyxl
import re
import csv
import json
from datetime import datetime

SRC = "/mnt/user-data/uploads/serie_historica_mai26.xlsx"
HEADER_ROW = 5

# metadata: sheet_name -> (periodicidade, unidade_tipo)
SHEET_META = {
    '1.1': ('mensal', 'corrente'), '1.1-A': ('mensal', 'constante_ipca'),
    '1.2': ('mensal', 'corrente'), '1.2-A': ('mensal', 'constante_ipca'),
    '1.2-B': ('mensal', 'constante_ipca_acum12m'),
    '1.3': ('mensal', 'corrente'), '1.3-A': ('mensal', 'constante_ipca'),
    '1.4': ('mensal', 'corrente'), '1.4-A': ('mensal', 'constante_ipca'),
    '1.5': ('mensal', 'corrente'), '1.5-A': ('mensal', 'constante_ipca'),
    '1.6': ('mensal', 'corrente'),
    '2.1': ('anual', 'corrente'), '2.1-A': ('anual', 'pct_pib'),
    '2.2': ('anual', 'corrente'), '2.2-A': ('anual', 'pct_pib'),
    '2.3': ('anual', 'corrente'), '2.3-A': ('anual', 'pct_pib'),
    '2.4': ('anual', 'corrente'), '2.4-A': ('anual', 'pct_pib'),
    '2.5': ('anual', 'corrente'), '2.5-A': ('anual', 'pct_pib'),
    '4.1': ('trimestral', 'corrente'),
    '4.2': ('trimestral', 'corrente'),
}
# 3.1 / 3.2 sao tabelas "comparativas" (mes vs mes do ano anterior, acumulados, variacoes)
# derivadas dos dados de 1.1/1.2 -> tratadas a parte (nao entram na serie longa principal)
SKIP_SHEETS = {'Índice', '3.1', '3.2'}

def col_letter(n):
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)

def parse_sheet(ws, sheet_name, periodicidade, unidade_tipo):
    titulo = ws.cell(row=2, column=1).value
    unidade_desc = ws.cell(row=3, column=1).value

    header_vals = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
    # detect two-column (codigo, label) format: header col B is None AND first data row col A is int
    first_data_row = HEADER_ROW + 1
    col_a_first = ws.cell(row=first_data_row, column=1).value
    two_col_format = isinstance(col_a_first, int) and header_vals[1] is None

    if two_col_format:
        label_col = 2
        data_start_col = 3
    else:
        label_col = 1
        data_start_col = 2

    # periodos: from data_start_col to last non-None header cell
    periodos = []  # list of (col_idx, periodo_str)
    for c in range(data_start_col, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v is None:
            break
        if isinstance(v, datetime):
            periodo = v.strftime('%Y-%m')
        else:
            periodo = str(v)
        periodos.append((c, periodo))

    rows = []
    r = first_data_row
    while r <= ws.max_row:
        label = ws.cell(row=r, column=label_col).value
        if label is None:
            break  # fim do bloco de dados (comeco das notas de rodape)
        codigo_raw = ws.cell(row=r, column=1).value if two_col_format else None
        discriminacao = str(label).strip()
        # extrai codigo hierarquico tipo "1.1.01" ou "4.3.20" do inicio do texto, se existir
        m = re.match(r'^([0-9]+(?:\.[0-9]+)*\.?)\s+(.*)', discriminacao)
        if m:
            codigo = m.group(1).rstrip('.')
            desc = m.group(2).strip()
        else:
            codigo = str(codigo_raw) if codigo_raw is not None else None
            desc = discriminacao
        nivel = codigo.count('.') + 1 if codigo else 1

        for c, periodo in periodos:
            valor = ws.cell(row=r, column=c).value
            if valor is None:
                continue
            rows.append({
                'tabela': sheet_name,
                'periodicidade': periodicidade,
                'unidade_tipo': unidade_tipo,
                'codigo': codigo,
                'descricao': desc,
                'nivel': nivel,
                'periodo': periodo,
                'valor': valor,
            })
        r += 1
    return titulo, unidade_desc, rows

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    all_rows = []
    tabelas_meta = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        if sheet_name not in SHEET_META:
            print('AVISO: sheet sem metadata, pulando ->', sheet_name)
            continue
        periodicidade, unidade_tipo = SHEET_META[sheet_name]
        ws = wb[sheet_name]
        titulo, unidade_desc, rows = parse_sheet(ws, sheet_name, periodicidade, unidade_tipo)
        tabelas_meta[sheet_name] = {'titulo': titulo, 'unidade_desc': unidade_desc, 'n_linhas': len(rows)}
        all_rows.extend(rows)
        print(f'{sheet_name}: {len(rows)} registros | {titulo}')

    print('\nTOTAL registros:', len(all_rows))

    # grava CSV (fonte de verdade, uma linha por tabela/rubrica/periodo)
    with open('rtn_long.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['tabela','periodicidade','unidade_tipo','codigo','descricao','nivel','periodo','valor'])
        writer.writeheader()
        writer.writerows(all_rows)

    with open('rtn_meta.json', 'w', encoding='utf-8') as f:
        json.dump(tabelas_meta, f, ensure_ascii=False, indent=2)

if __name__ == '__main__':
    main()
